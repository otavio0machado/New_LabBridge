"""
Utilitários de Exportação - PDF, CSV e Excel
Para exportação de análises salvas e relatórios
"""
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any, Optional
import csv
import json


def generate_analyses_excel(analyses: List[Dict[str, Any]]) -> bytes:
    """
    Gera arquivo Excel (.xlsx) com resumo das análises salvas
    
    Args:
        analyses: Lista de análises do banco de dados
        
    Returns:
        bytes: Conteúdo do arquivo Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl é necessário para exportar Excel. Instale com: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo de Análises"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="center")
    money_alignment = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Cabeçalhos
    headers = [
        "ID",
        "Nome da Análise",
        "Data da Análise",
        "Total COMPULAB (R$)",
        "Total SIMUS (R$)",
        "Diferença (R$)",
        "Pacientes COMPULAB",
        "Pacientes SIMUS",
        "Total Divergências",
        "Arquivo COMPULAB",
        "Arquivo SIMUS",
        "Criado em"
    ]
    
    # Aplicar cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Dados
    for row_num, analysis in enumerate(analyses, 2):
        row_data = [
            analysis.get("id", ""),
            analysis.get("analysis_name", ""),
            analysis.get("analysis_date", ""),
            analysis.get("compulab_total", 0) or 0,
            analysis.get("simus_total", 0) or 0,
            analysis.get("difference", 0) or 0,
            analysis.get("compulab_patients", 0) or 0,
            analysis.get("simus_patients", 0) or 0,
            analysis.get("total_divergences", 0) or 0,
            analysis.get("compulab_file_name", ""),
            analysis.get("simus_file_name", ""),
            analysis.get("created_at", "")[:19] if analysis.get("created_at") else "",
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            
            # Formatação de valores monetários
            if col_num in [4, 5, 6]:  # Colunas de dinheiro
                cell.number_format = '#,##0.00'
                cell.alignment = money_alignment
            else:
                cell.alignment = cell_alignment
    
    # Ajustar largura das colunas
    column_widths = [36, 30, 15, 18, 18, 15, 18, 15, 18, 25, 25, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Congelar primeira linha
    ws.freeze_panes = "A2"
    
    # === Aba de Totais ===
    ws_totals = wb.create_sheet("Totais")
    
    # Calcular totais
    total_compulab = sum(a.get("compulab_total", 0) or 0 for a in analyses)
    total_simus = sum(a.get("simus_total", 0) or 0 for a in analyses)
    total_diff = total_compulab - total_simus
    total_divergences = sum(a.get("total_divergences", 0) or 0 for a in analyses)
    
    totals_data = [
        ["Resumo Geral", ""],
        ["", ""],
        ["Total de Análises", len(analyses)],
        ["Total Faturamento COMPULAB", total_compulab],
        ["Total Faturamento SIMUS", total_simus],
        ["Diferença Total", total_diff],
        ["Total de Divergências", total_divergences],
        ["", ""],
        ["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
    ]
    
    for row_num, (label, value) in enumerate(totals_data, 1):
        ws_totals.cell(row=row_num, column=1, value=label).font = Font(bold=True if row_num in [1, 3, 4, 5, 6, 7, 9] else False)
        cell = ws_totals.cell(row=row_num, column=2, value=value)
        if row_num in [4, 5, 6]:
            cell.number_format = 'R$ #,##0.00'
    
    ws_totals.column_dimensions['A'].width = 30
    ws_totals.column_dimensions['B'].width = 25
    
    # Salvar em bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_analyses_csv(analyses: List[Dict[str, Any]]) -> str:
    """
    Gera arquivo CSV com resumo das análises salvas
    
    Args:
        analyses: Lista de análises do banco de dados
        
    Returns:
        str: Conteúdo do arquivo CSV
    """
    buffer = BytesIO()
    
    # Usar StringIO para CSV
    import io
    string_buffer = io.StringIO()
    
    headers = [
        "ID",
        "Nome da Análise",
        "Data da Análise",
        "Total COMPULAB (R$)",
        "Total SIMUS (R$)",
        "Diferença (R$)",
        "Pacientes COMPULAB",
        "Pacientes SIMUS",
        "Total Divergências",
        "Arquivo COMPULAB",
        "Arquivo SIMUS",
        "Criado em"
    ]
    
    writer = csv.writer(string_buffer, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    
    for analysis in analyses:
        row = [
            analysis.get("id", ""),
            analysis.get("analysis_name", ""),
            analysis.get("analysis_date", ""),
            f"{analysis.get('compulab_total', 0) or 0:.2f}".replace('.', ','),
            f"{analysis.get('simus_total', 0) or 0:.2f}".replace('.', ','),
            f"{analysis.get('difference', 0) or 0:.2f}".replace('.', ','),
            analysis.get("compulab_patients", 0) or 0,
            analysis.get("simus_patients", 0) or 0,
            analysis.get("total_divergences", 0) or 0,
            analysis.get("compulab_file_name", ""),
            analysis.get("simus_file_name", ""),
            analysis.get("created_at", "")[:19] if analysis.get("created_at") else "",
        ]
        writer.writerow(row)
    
    return string_buffer.getvalue()


def generate_analysis_detail_excel(analysis: Dict[str, Any], items: List[Dict[str, Any]]) -> bytes:
    """
    Gera Excel detalhado de uma análise específica com todos os itens
    
    Args:
        analysis: Dados da análise
        items: Lista de itens (divergências) da análise
        
    Returns:
        bytes: Conteúdo do arquivo Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl é necessário para exportar Excel")
    
    wb = openpyxl.Workbook()
    
    # === Aba Resumo ===
    ws_summary = wb.active
    ws_summary.title = "Resumo"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    
    summary_data = [
        ["Relatório de Análise", ""],
        ["", ""],
        ["Nome", analysis.get("analysis_name", "")],
        ["Data", analysis.get("analysis_date", "")],
        ["Descrição", analysis.get("description", "")],
        ["", ""],
        ["Faturamento COMPULAB", analysis.get("compulab_total", 0) or 0],
        ["Faturamento SIMUS", analysis.get("simus_total", 0) or 0],
        ["Diferença", analysis.get("difference", 0) or 0],
        ["", ""],
        ["Pacientes COMPULAB", analysis.get("compulab_patients", 0) or 0],
        ["Pacientes SIMUS", analysis.get("simus_patients", 0) or 0],
        ["Total Divergências", analysis.get("total_divergences", 0) or 0],
    ]
    
    for row_num, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_num, column=1, value=label).font = Font(bold=True)
        cell = ws_summary.cell(row=row_num, column=2, value=value)
        if row_num in [7, 8, 9]:
            cell.number_format = 'R$ #,##0.00'
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40
    
    # === Aba de Divergências ===
    if items:
        ws_items = wb.create_sheet("Divergências")
        
        headers = ["Tipo", "Paciente", "Exame", "Valor COMPULAB", "Valor SIMUS", "Diferença", "Anotação"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_items.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_num, item in enumerate(items, 2):
            item_data = item.get("item_data", {})
            if isinstance(item_data, str):
                try:
                    item_data = json.loads(item_data)
                except:
                    item_data = {}
            
            ws_items.cell(row=row_num, column=1, value=item.get("item_type", ""))
            ws_items.cell(row=row_num, column=2, value=item_data.get("patient", ""))
            ws_items.cell(row=row_num, column=3, value=item_data.get("exam_name", ""))
            ws_items.cell(row=row_num, column=4, value=item_data.get("compulab_value", 0) or 0).number_format = '#,##0.00'
            ws_items.cell(row=row_num, column=5, value=item_data.get("simus_value", 0) or 0).number_format = '#,##0.00'
            ws_items.cell(row=row_num, column=6, value=item_data.get("difference", 0) or 0).number_format = '#,##0.00'
            ws_items.cell(row=row_num, column=7, value=item.get("annotation", ""))
        
        # Ajustar larguras
        ws_items.column_dimensions['A'].width = 20
        ws_items.column_dimensions['B'].width = 30
        ws_items.column_dimensions['C'].width = 35
        ws_items.column_dimensions['D'].width = 18
        ws_items.column_dimensions['E'].width = 18
        ws_items.column_dimensions['F'].width = 15
        ws_items.column_dimensions['G'].width = 30
        
        ws_items.freeze_panes = "A2"
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_combined_pdf(analyses: List[Dict[str, Any]]) -> bytes:
    """
    Gera PDF combinado com resumo de todas as análises
    
    Args:
        analyses: Lista de análises
        
    Returns:
        bytes: Conteúdo do arquivo PDF
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor("#1E3A5F"),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    h2_style = ParagraphStyle(
        'H2',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor("#2563EB"),
        spaceBefore=15,
        spaceAfter=10
    )
    
    story = []
    
    # Título
    story.append(Paragraph("Relatório Consolidado de Análises", title_style))
    story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", 
                          ParagraphStyle('Date', parent=styles['Normal'], alignment=TA_CENTER, textColor=colors.grey)))
    story.append(Spacer(1, 1*cm))
    
    # Resumo Geral
    total_compulab = sum(a.get("compulab_total", 0) or 0 for a in analyses)
    total_simus = sum(a.get("simus_total", 0) or 0 for a in analyses)
    total_diff = total_compulab - total_simus
    
    story.append(Paragraph("Resumo Geral", h2_style))
    
    summary_data = [
        ["Métrica", "Valor"],
        ["Total de Análises", str(len(analyses))],
        ["Faturamento COMPULAB", f"R$ {total_compulab:,.2f}"],
        ["Faturamento SIMUS", f"R$ {total_simus:,.2f}"],
        ["Diferença Total", f"R$ {total_diff:,.2f}"],
    ]
    
    summary_table = Table(summary_data, colWidths=[8*cm, 6*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 1*cm))
    
    # Lista de Análises
    story.append(Paragraph("Detalhamento por Análise", h2_style))
    
    analysis_data = [["Nome", "Data", "COMPULAB", "SIMUS", "Diferença"]]
    for a in analyses:
        analysis_data.append([
            a.get("analysis_name", "")[:30],
            a.get("analysis_date", ""),
            f"R$ {a.get('compulab_total', 0) or 0:,.2f}",
            f"R$ {a.get('simus_total', 0) or 0:,.2f}",
            f"R$ {a.get('difference', 0) or 0:,.2f}",
        ])
    
    analysis_table = Table(analysis_data, colWidths=[5*cm, 2.5*cm, 3.5*cm, 3.5*cm, 3*cm])
    analysis_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    story.append(analysis_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
